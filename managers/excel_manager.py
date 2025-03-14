import os
import openpyxl
from openpyxl import Workbook

class ExcelManager:
    """
    Manager for all Excel-related operations including creating, updating, and formatting Excel files.
    """
    def __init__(self):
        # Default filename for Excel output
        self.default_excel_filename = "ui_defects.xlsx"
    
    def create_or_load_workbook(self, file_path):
        """Create a new workbook or load an existing one"""
        if os.path.exists(file_path):
            # Load existing workbook
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
        else:
            # Create a new workbook and set up the headers
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Defect Result"
            ws.append(["Filename", "Title", "Result", "Defect URL"])
        
        return wb, ws
    
    def save_defect_result(self, category_folder, filename, result_text):
        """
        Save defect result to Excel file
        
        Args:
            category_folder (str): Path to the category folder
            filename (str): The filename (with or without extension)
            result_text (str): Result text to save
            
        Returns:
            bool: True if successful, False otherwise
        """
        excel_path = os.path.join(category_folder, self.default_excel_filename)
        
        # Strip the file extension to save only the base filename
        base_filename, _ = os.path.splitext(filename)
        
        try:
            # Get workbook and worksheet
            wb, ws = self.create_or_load_workbook(excel_path)
            
            # Append data to the worksheet
            ws.append([base_filename, "", result_text, ""])
            
            # Save the workbook
            wb.save(excel_path)
            return True
        except Exception as e:
            print(f"Failed to update Excel file: {str(e)}")
            return False
    
    def generate_summary_report(self, base_folder):
        """
        Generate a summary report Excel file with statistics from all categories
        
        Args:
            base_folder (str): Base folder containing category folders
            
        Returns:
            str: Path to the generated report file or None if failed
        """
        # This is a placeholder for future functionality
        # You could implement summary statistics across all defects here
        pass
        
    def apply_formatting(self, workbook, worksheet):
        """
        Apply formatting to an Excel worksheet (colors, column widths, etc.)
        
        Args:
            workbook (Workbook): The openpyxl workbook
            worksheet: The worksheet to format
        """
        # This is a placeholder for future formatting functionality
        # For example, you could set column widths, add colors to headers, etc.
        pass 